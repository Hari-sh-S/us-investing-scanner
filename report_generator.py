"""
Report Generator Module
========================
Handles Excel and PDF export functionality for backtest results.
"""

import io
import pandas as pd
import numpy as np
from datetime import datetime


def create_excel_with_charts(config, metrics, engine=None):
    """
    Create an Excel file with all backtest data and embedded charts.
    
    Args:
        config: Backtest configuration dictionary
        metrics: Performance metrics dictionary
        engine: PortfolioEngine instance (optional, for detailed data)
    
    Returns:
        BytesIO object containing the Excel file
    """
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference, AreaChart
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = io.BytesIO()
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # ==================== SHEET 1: INPUTS ====================
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    
    inputs_data = [
        ["Parameter", "Value"],
        ["Strategy Name", config.get('name', 'Backtest')],
        ["Starting Capital", f"₹{config['initial_capital']:,}"],
        ["Universe", config['universe_name']],
        ["No. of Stocks in Portfolio", config['num_stocks']],
        ["Exit Rank", config.get('exit_rank', config['num_stocks'])],
        ["Rebalance Frequency", config['rebalance_freq']],
        ["Start Date", config['start_date']],
        ["End Date", config['end_date']],
        ["Regime Filter", str(config.get('regime_config', 'None'))],
        ["Uncorrelated Asset", str(config.get('uncorrelated_config', 'None'))],
        ["Scoring Formula", config['formula']]
    ]
    
    for row_idx, row in enumerate(inputs_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_inputs.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
    
    ws_inputs.column_dimensions['A'].width = 25
    ws_inputs.column_dimensions['B'].width = 50
    
    # ==================== SHEET 2: PERFORMANCE METRICS ====================
    ws_perf = wb.create_sheet("Performance Metrics")
    
    perf_data = [
        ["Metric", "Value"],
        ["Start Date", config['start_date']],
        ["End Date", config['end_date']],
        ["Invested Capital", f"₹{config['initial_capital']:,}"],
        ["Final Value", f"₹{metrics['Final Value']:,.2f}"],
        ["Total Return", f"₹{metrics['Total Return']:,.2f}"],
        ["Return %", f"{metrics['Return %']:.2f}%"],
        ["CAGR %", f"{metrics['CAGR %']:.2f}%"],
        ["Max Drawdown %", f"{metrics['Max Drawdown %']:.2f}%"],
        ["Volatility %", f"{metrics.get('Volatility %', 0):.2f}%"],
        ["Sharpe Ratio", f"{metrics['Sharpe Ratio']:.2f}"],
        ["Win Rate %", f"{metrics['Win Rate %']:.2f}%"],
        ["Total Trades", metrics['Total Trades']],
        ["Max Consecutive Wins", metrics.get('Max Consecutive Wins', 0)],
        ["Max Consecutive Losses", metrics.get('Max Consecutive Losses', 0)],
        ["Avg Win", f"₹{metrics.get('Avg Win', 0):,.2f}"],
        ["Avg Loss", f"₹{metrics.get('Avg Loss', 0):,.2f}"],
        ["Expectancy", f"₹{metrics.get('Expectancy', 0):,.2f}"],
        ["Days to Recover from DD", metrics.get('Days to Recover from DD', 0)],
        ["Trades to Recover from DD", metrics.get('Trades to Recover from DD', 0)],
        ["Total Turnover", f"₹{metrics.get('Total Turnover', 0):,.2f}"],
        ["Total Charges (Zerodha)", f"₹{metrics.get('Total Charges', 0):,.2f}"],
        ["STT/CTT", f"₹{metrics.get('STT/CTT', 0):,.2f}"],
        ["Transaction Charges", f"₹{metrics.get('Transaction Charges', 0):,.2f}"],
        ["SEBI Charges", f"₹{metrics.get('SEBI Charges', 0):,.2f}"],
        ["Stamp Charges", f"₹{metrics.get('Stamp Charges', 0):,.2f}"],
        ["GST", f"₹{metrics.get('GST', 0):,.2f}"],
    ]
    
    for row_idx, row in enumerate(perf_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_perf.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    ws_perf.column_dimensions['A'].width = 30
    ws_perf.column_dimensions['B'].width = 25
    
    # ==================== SHEET 3: MONTHLY RETURNS ====================
    ws_monthly = wb.create_sheet("Monthly Returns")
    
    if engine and hasattr(engine, 'portfolio_df') and not engine.portfolio_df.empty:
        monthly_returns = engine.get_monthly_returns()
        if not monthly_returns.empty:
            # Write header
            headers = ['Year'] + list(monthly_returns.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_monthly.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            # Write data
            for row_idx, (year, row) in enumerate(monthly_returns.iterrows(), 2):
                ws_monthly.cell(row=row_idx, column=1, value=year).border = thin_border
                for col_idx, value in enumerate(row, 2):
                    cell = ws_monthly.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if pd.notna(value):
                        cell.value = round(value, 2)
                        if value > 0:
                            cell.fill = positive_fill
                        elif value < 0:
                            cell.fill = negative_fill
    
    # ==================== SHEET 4: EQUITY CURVE CHART ====================
    ws_chart = wb.create_sheet("Charts")
    
    if engine and hasattr(engine, 'portfolio_df') and not engine.portfolio_df.empty:
        # Prepare chart data
        chart_df = engine.portfolio_df[['Portfolio Value']].reset_index()
        chart_df.columns = ['Date', 'Portfolio Value']
        
        # Sample data for chart (max 500 points for performance)
        if len(chart_df) > 500:
            step = len(chart_df) // 500
            chart_df = chart_df.iloc[::step].copy()
        
        # Calculate drawdown for chart
        running_max = chart_df['Portfolio Value'].cummax()
        chart_df['Drawdown %'] = (chart_df['Portfolio Value'] - running_max) / running_max * 100
        
        # Write chart data
        ws_chart.cell(row=1, column=1, value="Date")
        ws_chart.cell(row=1, column=2, value="Portfolio Value")
        ws_chart.cell(row=1, column=3, value="Drawdown %")
        
        for row_idx, (_, row) in enumerate(chart_df.iterrows(), 2):
            ws_chart.cell(row=row_idx, column=1, value=row['Date'])
            ws_chart.cell(row=row_idx, column=2, value=row['Portfolio Value'])
            ws_chart.cell(row=row_idx, column=3, value=row['Drawdown %'])
        
        data_rows = len(chart_df) + 1
        
        # Create Equity Curve Chart
        chart1 = LineChart()
        chart1.title = "Equity Curve"
        chart1.style = 10
        chart1.y_axis.title = "Portfolio Value (₹)"
        chart1.x_axis.title = "Date"
        chart1.width = 20
        chart1.height = 10
        
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=data_rows)
        categories = Reference(ws_chart, min_col=1, min_row=2, max_row=data_rows)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.series[0].graphicalProperties.line.solidFill = "28A745"
        
        ws_chart.add_chart(chart1, "E2")
        
        # Create Drawdown Chart
        chart2 = AreaChart()
        chart2.title = "Drawdown Analysis"
        chart2.style = 10
        chart2.y_axis.title = "Drawdown %"
        chart2.x_axis.title = "Date"
        chart2.width = 20
        chart2.height = 10
        
        data2 = Reference(ws_chart, min_col=3, min_row=1, max_row=data_rows)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(categories)
        chart2.series[0].graphicalProperties.solidFill = "DC3545"
        chart2.series[0].graphicalProperties.line.solidFill = "DC3545"
        
        ws_chart.add_chart(chart2, "E22")
    
    # ==================== SHEET 5: DAILY REPORT ====================
    ws_daily = wb.create_sheet("Daily Report")
    
    if engine and hasattr(engine, 'portfolio_df') and not engine.portfolio_df.empty:
        daily_data = engine.portfolio_df.copy()
        daily_data = daily_data.reset_index()
        daily_data['Date'] = daily_data['Date'].dt.strftime('%Y-%m-%d')
        
        # Write headers
        headers = list(daily_data.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_daily.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Write data (limit to last 1000 rows for file size)
        data_to_write = daily_data.tail(1000)
        for row_idx, (_, row) in enumerate(data_to_write.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_daily.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
    
    # ==================== SHEET 6: TRADE HISTORY ====================
    ws_trades = wb.create_sheet("Trade History")
    
    if engine and hasattr(engine, 'trades_df') and not engine.trades_df.empty:
        trades_df = engine.trades_df.copy()
        if 'Date' in trades_df.columns:
            trades_df['Date'] = pd.to_datetime(trades_df['Date']).dt.strftime('%Y-%m-%d')
        
        # Write headers
        headers = list(trades_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Write data
        for row_idx, (_, row) in enumerate(trades_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_trades.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Color BUY/SELL cells
                if col_idx == 3:  # Action column
                    if value == 'BUY':
                        cell.fill = positive_fill
                    elif value == 'SELL':
                        cell.fill = negative_fill
    
    wb.save(output)
    output.seek(0)
    return output


def create_pdf_report(config, metrics, engine=None):
    """
    Create a professional PDF report of backtest results.
    
    Args:
        config: Backtest configuration dictionary
        metrics: Performance metrics dictionary
        engine: PortfolioEngine instance (optional, for charts)
    
    Returns:
        BytesIO object containing the PDF file
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), 
                           leftMargin=1*cm, rightMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F4E79'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1F4E79'),
        spaceBefore=20,
        spaceAfter=10
    )
    
    # ==================== HEADER ====================
    story.append(Paragraph("BACKTEST REPORT", title_style))
    story.append(Paragraph(
        f"{config.get('name', 'Strategy')} | {config['start_date']} to {config['end_date']} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        subtitle_style
    ))
    
    # ==================== STRATEGY OVERVIEW ====================
    story.append(Paragraph("Strategy Overview", section_style))
    
    overview_data = [
        ["Universe", config['universe_name'], "Initial Capital", f"₹{config['initial_capital']:,}"],
        ["Portfolio Size", f"{config['num_stocks']} stocks", "Rebalance", config['rebalance_freq']],
        ["Exit Rank", config.get('exit_rank', config['num_stocks']), "Regime Filter", "Yes" if config.get('regime_config') else "No"],
    ]
    
    overview_table = Table(overview_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(overview_table)
    story.append(Spacer(1, 20))
    
    # ==================== KEY PERFORMANCE INDICATORS ====================
    story.append(Paragraph("Key Performance Indicators", section_style))
    
    # Format values with colors
    cagr_val = metrics['CAGR %']
    sharpe_val = metrics['Sharpe Ratio']
    dd_val = metrics['Max Drawdown %']
    
    kpi_data = [
        ["Final Value", "Total Return", "CAGR", "Max Drawdown"],
        [f"₹{metrics['Final Value']:,.0f}", f"₹{metrics['Total Return']:,.0f}", 
         f"{cagr_val:.2f}%", f"{dd_val:.2f}%"],
        ["Sharpe Ratio", "Win Rate", "Total Trades", "Expectancy"],
        [f"{sharpe_val:.2f}", f"{metrics['Win Rate %']:.1f}%", 
         f"{metrics['Total Trades']}", f"₹{metrics.get('Expectancy', 0):,.0f}"],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 2), (-1, 2), 10),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('FONTSIZE', (0, 3), (-1, 3), 14),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # ==================== ADDITIONAL METRICS ====================
    story.append(Paragraph("Trading Statistics", section_style))
    
    stats_data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Volatility", f"{metrics.get('Volatility %', 0):.2f}%", "Max Consecutive Wins", f"{metrics.get('Max Consecutive Wins', 0)}"],
        ["Avg Win", f"₹{metrics.get('Avg Win', 0):,.0f}", "Max Consecutive Losses", f"{metrics.get('Max Consecutive Losses', 0)}"],
        ["Avg Loss", f"₹{metrics.get('Avg Loss', 0):,.0f}", "Days to Recover from DD", f"{metrics.get('Days to Recover from DD', 0)}"],
        ["Total Turnover", f"₹{metrics.get('Total Turnover', 0):,.0f}", "Total Charges", f"₹{metrics.get('Total Charges', 0):,.0f}"],
    ]
    
    stats_table = Table(stats_data, colWidths=[4*cm, 4*cm, 4.5*cm, 4*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8F4FD')),
        ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#E8F4FD')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(stats_table)
    
    # ==================== EQUITY CURVE CHART ====================
    # Using matplotlib for PDF charts (works without Chrome/kaleido)
    if engine and hasattr(engine, 'portfolio_df') and not engine.portfolio_df.empty:
        story.append(PageBreak())
        story.append(Paragraph("Performance Charts", section_style))
        
        try:
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            
            # Equity Curve Chart
            fig1, ax1 = plt.subplots(figsize=(12, 5))
            dates = engine.portfolio_df.index
            values = engine.portfolio_df['Portfolio Value']
            
            ax1.fill_between(dates, values, alpha=0.3, color='#28A745')
            ax1.plot(dates, values, color='#28A745', linewidth=2)
            ax1.set_title('Equity Curve', fontsize=14, fontweight='bold', color='#1F4E79')
            ax1.set_xlabel('Date', fontsize=10)
            ax1.set_ylabel('Portfolio Value', fontsize=10)
            ax1.grid(True, alpha=0.3)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # Save to bytes
            equity_bytes = io.BytesIO()
            fig1.savefig(equity_bytes, format='png', dpi=150, bbox_inches='tight', 
                        facecolor='white', edgecolor='none')
            equity_bytes.seek(0)
            plt.close(fig1)
            
            equity_img = Image(equity_bytes, width=22*cm, height=9*cm)
            story.append(equity_img)
            story.append(Spacer(1, 20))
            
            # Drawdown Chart
            running_max = values.cummax()
            dd = (values - running_max) / running_max * 100
            
            fig2, ax2 = plt.subplots(figsize=(12, 4))
            ax2.fill_between(dates, dd, alpha=0.3, color='#DC3545')
            ax2.plot(dates, dd, color='#DC3545', linewidth=1.5)
            ax2.set_title('Drawdown Analysis', fontsize=14, fontweight='bold', color='#1F4E79')
            ax2.set_xlabel('Date', fontsize=10)
            ax2.set_ylabel('Drawdown %', fontsize=10)
            ax2.grid(True, alpha=0.3)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            dd_bytes = io.BytesIO()
            fig2.savefig(dd_bytes, format='png', dpi=150, bbox_inches='tight',
                        facecolor='white', edgecolor='none')
            dd_bytes.seek(0)
            plt.close(fig2)
            
            dd_img = Image(dd_bytes, width=22*cm, height=7*cm)
            story.append(dd_img)
            
        except Exception:
            # If matplotlib fails for any reason, just skip charts silently
            pass
    
    # ==================== MONTHLY RETURNS ====================
    if engine and hasattr(engine, 'get_monthly_returns'):
        monthly_returns = engine.get_monthly_returns()
        if not monthly_returns.empty:
            story.append(PageBreak())
            story.append(Paragraph("Monthly Returns (%)", section_style))
            
            # Build table data
            headers = ['Year'] + list(monthly_returns.columns)
            table_data = [headers]
            
            for year, row in monthly_returns.iterrows():
                row_data = [str(year)]
                for val in row:
                    if pd.notna(val):
                        row_data.append(f"{val:.1f}")
                    else:
                        row_data.append("-")
                table_data.append(row_data)
            
            # Create table
            col_widths = [1.5*cm] + [1.5*cm] * len(monthly_returns.columns)
            monthly_table = Table(table_data, colWidths=col_widths)
            
            # Style with conditional formatting
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Add conditional colors for positive/negative values
            for row_idx, row in enumerate(table_data[1:], 1):
                for col_idx, val in enumerate(row[1:], 1):
                    try:
                        num_val = float(val)
                        if num_val > 0:
                            style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#C6EFCE')))
                        elif num_val < 0:
                            style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FFC7CE')))
                    except ValueError:
                        pass
            
            monthly_table.setStyle(TableStyle(style_commands))
            story.append(monthly_table)
    
    # ==================== SCORING FORMULA ====================
    story.append(Spacer(1, 30))
    story.append(Paragraph("Scoring Formula", section_style))
    formula_style = ParagraphStyle(
        'Formula',
        parent=styles['Code'],
        fontSize=10,
        backColor=colors.HexColor('#F5F5F5'),
        borderPadding=10,
        leftIndent=20,
        rightIndent=20
    )
    story.append(Paragraph(config['formula'], formula_style))
    
    # ==================== FOOTER ====================
    story.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        f"Generated by Investing Scanner | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        footer_style
    ))
    
    doc.build(story)
    output.seek(0)
    return output


def prepare_complete_log_data(config, metrics, engine=None):
    """
    Prepare complete backtest data for log storage without truncation.
    
    Args:
        config: Backtest configuration
        metrics: Performance metrics
        engine: PortfolioEngine instance
    
    Returns:
        Dictionary with complete backtest data
    """
    log_data = {
        'config': config,
        'metrics': metrics,
        'portfolio_values': [],
        'trades': [],
        'monthly_returns': {}
    }
    
    if engine:
        # Store portfolio values (date + value pairs)
        if hasattr(engine, 'portfolio_df') and not engine.portfolio_df.empty:
            for date, row in engine.portfolio_df.iterrows():
                log_data['portfolio_values'].append({
                    'date': date.strftime('%Y-%m-%d'),
                    'value': float(row['Portfolio Value']),
                    'cash': float(row.get('Cash', 0)),
                    'positions': int(row.get('Positions', 0))
                })
        
        # Store all trades
        if hasattr(engine, 'trades_df') and not engine.trades_df.empty:
            for _, trade in engine.trades_df.iterrows():
                trade_dict = {}
                for col in trade.index:
                    val = trade[col]
                    if pd.isna(val):
                        trade_dict[col] = None
                    elif isinstance(val, (pd.Timestamp, datetime)):
                        trade_dict[col] = val.strftime('%Y-%m-%d')
                    elif isinstance(val, (np.integer, np.floating)):
                        trade_dict[col] = float(val)
                    else:
                        trade_dict[col] = val
                log_data['trades'].append(trade_dict)
        
        # Store monthly returns
        if hasattr(engine, 'get_monthly_returns'):
            monthly_df = engine.get_monthly_returns()
            if not monthly_df.empty:
                for year, row in monthly_df.iterrows():
                    year_data = {}
                    for month, val in row.items():
                        if pd.notna(val):
                            year_data[month] = float(val)
                        else:
                            year_data[month] = None
                    log_data['monthly_returns'][str(year)] = year_data
    
    return log_data
